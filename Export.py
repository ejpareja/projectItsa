+from tkinter import *
from tkinter import StringVar
from tkinter import messagebox
from tkinter import ttk
import openpyxl
import Queries
from datetime import date


def export_excel():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "reporte"
    
def get_users():
        name = i[1]
        role = i[2]
        sheet.append((ide,name, role))
    
    now = date.today()
   
    wb.save("reporteusuario"+str(now)+".xlsx")

def get_invoices()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "reporte"
    data = Queries.get_invoices()
    sheet.append(('id', 'nombre', 'rol'))
    for i in data:
        ide = i[0]
        name = i[1]
        role = i[2]
        sheet.append((ide,name, role))
    now = date.today()
    wb.save("reporteproductos"+str(now)+".xlsx")
export_excel()

def get_invoices()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "reporte"
    data = Queries.get_invoices()
    sheet.append(('id', 'nombre', 'rol'))
    for i in data:
        ide = i[0]
        name = i[1]
        role = i[2]
        sheet.append((ide,name, role))
    now = date.today()
    wb.save("reporteproductos"+str(now)+".xlsx")
def back():
    root.destroy()
    import menu.py
    menu.meMain(data)

font = ('Bahnschrift Light', 12) 
root = Tk()
root.geometry("760x400+300+100")
root.resizable(False,False)
root.title("Registrar Usuario")
root.iconbitmap('presupuesto.ico')
root.config(bg="#222f3e")
#Treeview
tree = ttk.Treeview(root)
tree["columns"] = ("Identificación", "Nombre", "Rol")
tree.column("#0", width=30, minwidth=270, stretch=NO)
tree.column("Identificación", width=130, minwidth=270, stretch=NO)
tree.column("Nombre", width=200, minwidth=270, stretch=NO)
tree.column("Rol", width=100, minwidth=270, stretch=NO)

tree.heading("#0",text="",anchor=W)
tree.heading("Identificación", text="Identificación",anchor=W)
tree.heading("Nombre", text="Nombre",anchor=W)
tree.heading("Rol", text="Rol",anchor=W)
tree.place(x=150,y=50)

#Button
Button(root, text="Exportar", font=font, bg="#c8d6e5", fg="black", command=export_excel).place(x=540, y=300)
Button(root, text="Regresar", font=font, bg="#c8d6e5", fg="black", command=back).place(x=150, y=300)
root.mainloop()